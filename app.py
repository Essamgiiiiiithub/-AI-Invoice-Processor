import streamlit as st
import os
import sqlite3
import subprocess
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
from ocr_engine import extract_text
from ai_extractor import extract_document_data
from data_handler import init_database, save_invoice, get_all_invoices, get_invoices_df, delete_invoice
from batch_processor import DocumentProcessor

import tempfile
DB_PATH = os.path.join(tempfile.gettempdir(), "documents.db")
EXCEL_PATH = os.path.join(os.getcwd(), "documents.xlsx")

# إعداد الصفحة
st.set_page_config(
    page_title="AI Document Processor",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# CSS للتصميم
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap');

*, *::before, *::after { font-family: 'Plus Jakarta Sans', sans-serif !important; box-sizing: border-box; }

footer, #MainMenu, header { visibility: hidden; }
.block-container { padding-top: 1.5rem !important; }

.hero {
    background: linear-gradient(135deg, #0f172a 0%, #1e3a5f 50%, #0f2942 100%);
    padding: 2rem 2.5rem;
    border-radius: 20px;
    color: white;
    margin-bottom: 1.5rem;
    position: relative;
    overflow: hidden;
}
.hero::before {
    content: '';
    position: absolute;
    top: -40px; right: -40px;
    width: 200px; height: 200px;
    background: radial-gradient(circle, rgba(99,179,237,0.15) 0%, transparent 70%);
    border-radius: 50%;
}
.hero::after {
    content: '';
    position: absolute;
    bottom: -30px; left: 30%;
    width: 300px; height: 150px;
    background: radial-gradient(ellipse, rgba(139,92,246,0.1) 0%, transparent 70%);
}
.hero-inner { position: relative; z-index: 1; }
.hero h1 { font-size: 1.9rem; font-weight: 800; margin: 0; letter-spacing: -0.5px; }
.hero p { font-size: 0.92rem; opacity: 0.7; margin: 0.4rem 0 0; }
.hero-badge {
    display: inline-block;
    background: rgba(99,179,237,0.2);
    border: 1px solid rgba(99,179,237,0.35);
    color: #63b3ed;
    padding: 3px 12px;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 600;
    margin-bottom: 0.75rem;
    letter-spacing: 0.5px;
}

.kpi-card {
    background: white;
    border-radius: 16px;
    padding: 1.2rem 1.4rem;
    border: 1px solid #e8edf5;
    box-shadow: 0 2px 8px rgba(15,23,42,0.06);
    position: relative;
    overflow: hidden;
    transition: transform .2s, box-shadow .2s;
}
.kpi-card:hover { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(15,23,42,0.1); }
.kpi-card::after {
    content: '';
    position: absolute;
    top: 0; right: 0;
    width: 4px; height: 100%;
    border-radius: 0 16px 16px 0;
}
.kpi-blue::after { background: #3b82f6; }
.kpi-purple::after { background: #8b5cf6; }
.kpi-green::after { background: #10b981; }
.kpi-orange::after { background: #f59e0b; }
.kpi-icon  { font-size: 1.6rem; margin-bottom: 0.5rem; }
.kpi-num   { font-size: 1.9rem; font-weight: 800; color: #0f172a; line-height: 1; }
.kpi-label { font-size: 0.78rem; color: #64748b; font-weight: 500; margin-top: 4px; text-transform: uppercase; letter-spacing: 0.5px; }
.kpi-sub   { font-size: 0.78rem; color: #94a3b8; margin-top: 6px; }

/* ── Invoice Cards ── */
.inv-card {
    background: white;
    border-radius: 14px;
    padding: 1rem 1.2rem;
    border: 1px solid #e8edf5;
    margin-bottom: 10px;
    box-shadow: 0 1px 4px rgba(15,23,42,0.05);
    transition: box-shadow .2s;
}
.inv-card:hover { box-shadow: 0 4px 12px rgba(15,23,42,0.09); }
.inv-name  { font-size: 0.95rem; font-weight: 700; color: #0f172a; }
.inv-row   { font-size: 0.82rem; color: #475569; margin: 3px 0; }
.inv-amount { font-size: 1.1rem; font-weight: 700; color: #3b82f6; }
.badge {
    display: inline-block;
    padding: 2px 10px;
    border-radius: 20px;
    font-size: 0.73rem;
    font-weight: 600;
}
.badge-blue   { background: #dbeafe; color: #1d4ed8; }
.badge-green  { background: #dcfce7; color: #166534; }
.badge-purple { background: #ede9fe; color: #6d28d9; }

/* ── Upload zone ── */
.stFileUploader label {
    font-weight: 700 !important;
    color: #111827 !important;
    font-size: 1rem !important;
}
.uploadedFile,
[data-testid="stFileUploader"],
[data-testid="stFileUploader"] * {
    color: #111827 !important;
    font-weight: 700 !important;
    opacity: 1 !important;
}
[data-testid="stFileUploader"] {
    background: #f0f9ff !important;
    border: 2px dashed #3b82f6 !important;
    border-radius: 12px !important;
    padding: 0.5rem !important;
}
[data-testid="stFileUploader"]:hover {
    border-color: #1d4ed8 !important;
    background: #eff6ff !important;
}
section[data-testid="stFileUploadDropzone"],
section[data-testid="stFileUploadDropzone"] * {
    color: #111827 !important;
    font-weight: 700 !important;
    font-size: 0.9rem !important;
}
section[data-testid="stFileUploadDropzone"] {
    padding: 0.8rem !important;
    background: transparent !important;
}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    background: #eff6ff;
    border-radius: 14px;
    padding: 6px;
    gap: 10px;
    border: none !important;
    margin-bottom: 1rem;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 12px !important;
    font-weight: 800 !important;
    font-size: 1rem !important;
    letter-spacing: 0.02em !important;
    line-height: 1.4 !important;
    padding: 12px 24px !important;
    color: #111827 !important;
    border: none !important;
    background: #ffffff !important;
    text-shadow: none !important;
    transition: transform .2s ease, box-shadow .2s ease, color .2s ease, background .2s ease;
}
.stTabs [data-baseweb="tab"]:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 12px rgba(15,23,42,0.08) !important;
}
.stTabs [aria-selected="true"] {
    background: #ffffff !important;
    border: 1px solid #c7d2fe !important;
    color: #111827 !important;
    font-weight: 900 !important;
    box-shadow: 0 6px 16px rgba(15,23,42,0.08) !important;
}

/* ── Section titles ── */
.section-title {
    font-size: 1.05rem;
    font-weight: 700;
    color: #0f172a;
    margin-bottom: 1rem;
    display: flex;
    align-items: center;
    gap: 8px;
}
.section-title::after {
    content: '';
    flex: 1;
    height: 1px;
    background: #e2e8f0;
    margin-left: 8px;
}

/* ── Chart container ── */
.chart-box {
    background: white;
    border-radius: 16px;
    border: 1px solid #e8edf5;
    box-shadow: 0 2px 8px rgba(15,23,42,0.05);
    padding: 1rem;
    margin-bottom: 1rem;
}

/* ── Empty state ── */
.empty-state {
    text-align: center;
    padding: 3rem 1rem;
    color: #94a3b8;
}
.empty-state .emoji { font-size: 3rem; margin-bottom: 1rem; }
.empty-state p { font-size: 0.9rem; }

/* ── Expander (Danger Zone) ── */
.stExpander {
    border: 3px solid #ef4444 !important;
    border-radius: 12px !important;
    background: white !important;
}
.stExpander > button {
    background: linear-gradient(135deg, #fee2e2 0%, #fecaca 100%) !important;
    color: #991b1b !important;
    font-weight: 800 !important;
    font-size: 1.05rem !important;
    padding: 1.2rem !important;
    border: none !important;
}
.stExpander > button:hover {
    background: linear-gradient(135deg, #fecaca 0%, #fca5a5 100%) !important;
    box-shadow: 0 4px 12px rgba(239, 68, 68, 0.3) !important;
}
section[data-testid="stExpander"] {
    background: #fffbfb !important;
    border-radius: 12px !important;
}
section[data-testid="stExpander"] > div {
    padding: 1.5rem !important;
    background: #fffbfb !important;
}
section[data-testid="stExpander"] p,
section[data-testid="stExpander"] span,
section[data-testid="stExpander"] div {
    color: #0f172a !important;
    font-weight: 500 !important;
}


/* ── Warning boxes ── */
.stAlert {
    color: #0f172a !important;
    font-weight: 600 !important;
}
</style>
""", unsafe_allow_html=True)

init_database()
os.makedirs("outputs", exist_ok=True)

def update_excel(data, file_name):
    new_row = {
        "Document Type":   data.get("document_type") or "—",
        "Title":           data.get("title") or "—",
        "Description":     data.get("description") or "—",
        "Supplier Name":   data.get("supplier_name") or "—",
        "Document Number":  data.get("invoice_number") or "—",
        "Date":            data.get("date") or "—",
        "Total Amount":    data.get("total_amount") or "—",
        "Numeric Total":   data.get("numeric_total") or "",
        "Tax Amount":      data.get("tax_amount") or "—",
        "Currency":        data.get("currency") or "—",
        "Items":           ", ".join(data.get("items") or []) or "—",
        "Payment Method":  data.get("payment_method") or "—",
        "Notes":           data.get("notes") or "",
        "File Name":       file_name,
        "Added At":        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    try:
        if os.path.exists(EXCEL_PATH):
            df = pd.read_excel(EXCEL_PATH)
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        else:
            df = pd.DataFrame([new_row])

        with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Documents")
            ws = writer.sheets["Documents"]
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            hf = PatternFill("solid", fgColor="0F2942")
            hfont = Font(bold=True, color="FFFFFF", size=11)
            border = Border(left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"), bottom=Side(style="thin"))
            for cell in ws[1]:
                cell.fill = hf; cell.font = hfont
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.row % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor="F0F6FF")
            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 30
        return True
    except Exception as e:
        st.error(f"Excel error: {e}")
        return False

def make_chart(fig):
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font_family="Plus Jakarta Sans",
        margin=dict(l=10, r=10, t=40, b=10),
        legend=dict(bgcolor="rgba(0,0,0,0)"),
    )
    fig.update_xaxes(showgrid=False, zeroline=False)
    fig.update_yaxes(gridcolor="#f1f5f9", zeroline=False)
    return fig

# Hero
st.markdown("""
<div class="hero">
  <div class="hero-inner">
    <h2>AI Document Processor</h2>
    <p>Scan · Extract · Analyze · Export — all in seconds</p>
  </div>
</div>
""", unsafe_allow_html=True)

# KPI Row
df_all = get_invoices_df()
total_invoices = len(df_all)
total_suppliers = df_all["supplier_name"].nunique() if not df_all.empty else 0
total_amount = df_all["numeric_total"].sum() if not df_all.empty and "numeric_total" in df_all.columns else 0
avg_amount = df_all["numeric_total"].mean() if not df_all.empty and "numeric_total" in df_all.columns and total_invoices > 0 else 0

k1, k2, k3, k4 = st.columns(4)
with k1:
    st.markdown(f"""<div class="kpi-card kpi-blue">
        <div class="kpi-icon"></div>
        <div class="kpi-num">{total_invoices}</div>
        <div class="kpi-label">Total Documents</div>
    </div>""", unsafe_allow_html=True)
with k2:
    st.markdown(f"""<div class="kpi-card kpi-purple">
        <div class="kpi-num">{total_suppliers}</div>
        <div class="kpi-label">Suppliers</div>
    </div>""", unsafe_allow_html=True)
with k3:
    fmt = f"{total_amount:,.0f}" if total_amount else "—"
    st.markdown(f"""<div class="kpi-card kpi-green">
        <div class="kpi-num">{fmt}</div>
        <div class="kpi-label">Total Spent</div>
    </div>""", unsafe_allow_html=True)
with k4:
    fmt2 = f"{avg_amount:,.0f}" if avg_amount else "—"
    st.markdown(f"""<div class="kpi-card kpi-orange">
        <div class="kpi-num">{fmt2}</div>
        <div class="kpi-label">Avg per Document</div>
    </div>""", unsafe_allow_html=True)
tab1, tab2, tab3 = st.tabs(["Upload Document", "Dashboard", "History"])

with tab1:
    col_up, col_res = st.columns([1, 1], gap="large")

    with col_up:
        st.markdown('<div class="section-title">Upload New Document</div>', unsafe_allow_html=True)
        uploaded_files = st.file_uploader(
            "Click to browse or drag & drop files",
            type=["jpg", "jpeg", "png", "bmp", "pdf"],
            accept_multiple_files=True,
            label_visibility="visible",
        )
        if uploaded_files:
            def is_image(filename):
                return os.path.splitext(filename)[1].lower() in [".jpg", ".jpeg", ".png", ".bmp"]

            if len(uploaded_files) == 1:
                if is_image(uploaded_files[0].name):
                    st.image(uploaded_files[0], use_container_width=True, caption=uploaded_files[0].name)
                else:
                    st.markdown(f"**Uploaded document:** {uploaded_files[0].name}")
            else:
                st.markdown("**Selected documents:**")
                for file_item in uploaded_files:
                    if is_image(file_item.name):
                        st.write(f"- {file_item.name} (image)")
                    else:
                        st.write(f"- {file_item.name} (document)")
            st.markdown("<br>", unsafe_allow_html=True)

            if st.button("Analyze Documents", type="primary", use_container_width=True):
                if len(uploaded_files) > 1:
                    # استخدم معالج دفعي للملفات المتعددة (أسرع)
                    with st.spinner(f"⏳ Processing {len(uploaded_files)} documents with optimized batch processor..."):
                        processor = DocumentProcessor(max_workers=3)
                        results, errors = processor.process_files(uploaded_files)
                        
                        # حفظ النتائج دفعة واحدة
                        if results:
                            from data_handler import save_invoices_batch
                            data_list = [data for data, _ in results]
                            save_invoices_batch(data_list)
                            
                            # تحديث Excel
                            for data, filename in results:
                                update_excel(data, filename)
                        
                        # عرض النتائج
                        if errors:
                            st.warning(f"{len(errors)} document(s) had errors")
                            for filename, error in errors:
                                st.error(f"{filename}: {error}")
                        
                        if results:
                            st.success(f"{len(results)} document(s) analyzed and saved!")
                            st.session_state["last_result"] = results[0][0]
                        
                        st.rerun()
                else:
                    # معالجة ملف واحد بشكل طبيعي
                    uploaded_file = uploaded_files[0]
                    with st.spinner(f"⏳ Reading text from {uploaded_file.name}..."):
                        file_path = os.path.join(tempfile.gettempdir(), uploaded_file.name)
                        with open(file_path, "wb") as f:
                            f.write(uploaded_file.getbuffer())
                        text, error = extract_text(file_path)

                    if error:
                        st.error(f"{uploaded_file.name}: {error}")
                    elif not text:
                        st.error(f"{uploaded_file.name}: No text found. Try a clearer photo.")
                    else:
                        with st.spinner(f"🤖 AI extracting document data for {uploaded_file.name}..."):
                            data = extract_document_data(text)
                            save_invoice(data, uploaded_file.name)
                            excel_ok = update_excel(data, uploaded_file.name)
                            
                            if excel_ok:
                                st.success(f"{uploaded_file.name} analyzed and saved!")
                            else:
                                st.warning(f"Saved to database but Excel update failed.")
                            
                            st.session_state["last_result"] = data
                            st.rerun()

    with col_res:
        st.markdown('<div class="section-title">Extracted Information</div>', unsafe_allow_html=True)
        result = st.session_state.get("last_result")
        if result:
            # Amount highlight
            if result.get("total_amount"):
                st.metric("Total Amount", result.get("total_amount") or "—")

            r1, r2 = st.columns(2)
            with r1:
                st.metric("Supplier",    result.get("supplier_name") or "—")
                st.metric("Date",        result.get("date") or "—")
                st.metric("Tax",         result.get("tax_amount") or "—")
            with r2:
                st.metric("Document No.", result.get("invoice_number") or "—")
                st.metric("Payment",     result.get("payment_method") or "—")
                items = result.get("items") or []
                st.metric("Items",       len(items))

            if result.get("document_type"):
                st.metric("Document Type", result.get("document_type"))
            if result.get("title"):
                st.metric("Title", result.get("title"))
            if result.get("description"):
                st.metric("📝 Description", result.get("description"))

            if items:
                st.markdown("**Items / Services:**")
                for item in items:
                    st.markdown(f"<span class='badge badge-blue'>• {item}</span> ", unsafe_allow_html=True)

            if result.get("notes"):
                st.info(f" {result['notes']}")
        else:
            st.markdown('<div class="empty-state"><p>Upload and analyze a document to see extracted information here</p></div>', unsafe_allow_html=True)

# TAB 2 — DASHBOARD (Unified View)
with tab2:
    if df_all.empty:
        st.markdown('<div class="empty-state" style="padding:4rem"><p>No data yet - upload some documents first<br>and the dashboard will come alive!</p></div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="section-title">All Invoices Dashboard</div>', unsafe_allow_html=True)
        
        # إنشاء جدول موحد بكل المعلومات
        df_dashboard = df_all.copy()
        
        # تنسيق البيانات للعرض
        df_display = pd.DataFrame({
            'Supplier': df_dashboard['supplier_name'].fillna('—'),
            'Invoice #': df_dashboard['invoice_number'].fillna('—'),
            'Date': df_dashboard['date'].fillna('—'),
            'Type': df_dashboard['document_type'].fillna('Unknown'),
            'Total Amount': df_dashboard['total_amount'].fillna('—'),
            'Numeric': df_dashboard['numeric_total'].fillna(0),
            'Tax': df_dashboard['tax_amount'].fillna('—'),
            'Payment': df_dashboard['payment_method'].fillna('—'),
            'Title': df_dashboard['title'].fillna('—'),
            'Added At': df_dashboard['created_at'].fillna('—')
        })
        
        # إضافة خيارات الفرز والتصفية
        col_filter1, col_filter2, col_filter3 = st.columns(3)
        
        with col_filter1:
            supplier_filter = st.multiselect(
                "Filter by Supplier",
                options=['All'] + sorted(df_dashboard['supplier_name'].dropna().unique().tolist()),
                default=['All'],
                key="supplier_filter"
            )
        
        with col_filter2:
            type_filter = st.multiselect(
                "Filter by Document Type",
                options=['All'] + sorted(df_dashboard['document_type'].fillna('Unknown').unique().tolist()),
                default=['All'],
                key="type_filter"
            )
        
        with col_filter3:
            sort_by = st.selectbox(
                "Sort by",
                ["Latest", "Oldest", "Highest Amount", "Lowest Amount", "Supplier A-Z"],
                key="sort_by"
            )
        
        # تطبيق التصفية
        df_filtered = df_display.copy()
        
        if 'All' not in supplier_filter:
            df_filtered = df_filtered[df_filtered['Supplier'].isin(supplier_filter)]
        
        if 'All' not in type_filter:
            df_filtered = df_filtered[df_filtered['Type'].isin(type_filter)]
        
        # تطبيق الترتيب
        if sort_by == "Highest Amount":
            df_filtered = df_filtered.sort_values('Numeric', ascending=False)
        elif sort_by == "Lowest Amount":
            df_filtered = df_filtered.sort_values('Numeric', ascending=True)
        elif sort_by == "Oldest":
            df_filtered = df_filtered.sort_values('Date', ascending=True)
        elif sort_by == "Supplier A-Z":
            df_filtered = df_filtered.sort_values('Supplier', ascending=True)
        else:  # Latest (default)
            df_filtered = df_filtered.sort_values('Added At', ascending=False)
        
        # عرض الإحصائيات العامة
        summary_col1, summary_col2, summary_col3, summary_col4 = st.columns(4)
        
        with summary_col1:
            st.metric("Total Documents", len(df_filtered))
        
        with summary_col2:
            total_sum = df_filtered['Numeric'].sum()
            st.metric("Total Amount", f"{total_sum:,.0f}" if total_sum else "0")
        
        with summary_col3:
            total_tax = df_filtered[df_filtered['Tax'] != '—']['Tax'].apply(
                lambda x: float(x.split()[0].replace(',', '')) if isinstance(x, str) else 0
            ).sum()
            st.metric("Total Tax", f"{total_tax:,.0f}" if total_tax else "0")
        
        with summary_col4:
            avg_amount = df_filtered['Numeric'].mean()
            st.metric("Average Amount", f"{avg_amount:,.0f}" if avg_amount else "0")
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # عرض الجدول بشكل تفاعلي
        st.markdown('<div class="section-title">Complete Invoices Table</div>', unsafe_allow_html=True)
        
        # اختيار الأعمدة المراد عرضها
        col_select_area = st.expander("Customize Columns", expanded=False)
        with col_select_area:
            columns_to_show = st.multiselect(
                "Select columns to display",
                options=df_display.columns.tolist(),
                default=df_display.columns.tolist()[:8],
                key="columns_select"
            )
        
        df_final = df_filtered[columns_to_show]
        
        # عرض الجدول
        st.dataframe(
            df_final,
            use_container_width=True,
            height=500,
            hide_index=True,
            column_config={
                'Numeric': st.column_config.NumberColumn(format="%.0f"),
            }
        )
        
        st.markdown("<br>", unsafe_allow_html=True)

# TAB 3 — HISTORY
with tab3:
    col_list, col_actions = st.columns([2, 1], gap="large")

    with col_list:
        st.markdown('<div class="section-title">All Documents</div>', unsafe_allow_html=True)
        invoices = get_all_invoices()
        if invoices:
            for inv in invoices:
                amount_html = f'<span class="inv-amount">{inv.get("total_amount") or "—"}</span>' if inv.get("total_amount") else '<span style="color:#94a3b8">No amount</span>'
                st.write("test")
        else:
            st.markdown('<div class="empty-state"><p>No documents yet<br>Upload your first document!</p></div>', unsafe_allow_html=True)

    with col_actions:
        st.markdown('<div class="section-title">Actions</div>', unsafe_allow_html=True)

        # Export Excel
        if st.button("Export to Excel", use_container_width=True, type="primary"):
            if os.path.exists(EXCEL_PATH):
                subprocess.Popen(f'start excel "{EXCEL_PATH}"', shell=True)
                st.success("Opening Excel...")
            else:
                st.warning("No Excel file yet — analyze a document first!")

        # Download Excel
        if os.path.exists(EXCEL_PATH):
            with open(EXCEL_PATH, "rb") as f:
                st.download_button(
                    "Download Excel",
                    data=f,
                    file_name="documents.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        st.markdown("<br>", unsafe_allow_html=True)

        # Stats summary
        if not df_all.empty:
            st.markdown('<div style="background:#f8faff;border-radius:12px;padding:1rem;border:1px solid #e0e7ff;">', unsafe_allow_html=True)
            st.markdown("**Quick Stats**")
            if "numeric_total" in df_all.columns:
                valid = df_all["numeric_total"].dropna()
                if not valid.empty:
                    st.markdown(f"• Max: **{valid.max():,.0f}**")
                    st.markdown(f"• Min: **{valid.min():,.0f}**")
                    st.markdown(f"• Avg: **{valid.mean():,.0f}**")
            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        if st.button("Clear All Data", use_container_width=True, type="secondary"):
            conn = sqlite3.connect(DB_PATH)
            conn.execute("DELETE FROM invoices")
            conn.commit()
            conn.close()
            if os.path.exists(EXCEL_PATH):
                try:
                    os.remove(EXCEL_PATH)
                except PermissionError:
                    st.error("Close Excel first, then try again!")
                    st.stop()
            st.session_state.pop("last_result", None)
            st.success("All data cleared!")
            st.rerun()
